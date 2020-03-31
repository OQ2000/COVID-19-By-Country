import os
import os.path
import sys
from time import sleep
import webbrowser
import datetime
from datetime import time, date, timedelta
from openpyxl import Workbook, load_workbook, styles
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import xlrd 
import string
clear = lambda: os.system('cls')
clear()
totalFormatted = 0

def GetAllFiles():
    AllFiles = os.listdir(r"C:\Users\gunsl_000\Desktop\COVID-19 By Country\COVID-19-By-Country\Datasets\SeperateCountryFiles")
    print("Found",len(AllFiles),"Files")
    print("Opening And Saving All")
    sleep(2)
    return AllFiles

def Main(AllFiles):
    CountryFolderFilePath = r"C:\Users\gunsl_000\Desktop\COVID-19 By Country\COVID-19-By-Country\Datasets\SeperateCountryFiles/"
    filelen = len(AllFiles)
    for f in range(filelen):
        currentCountry =AllFiles[f]
        print("Saving",currentCountry)
        wb1 = load_workbook(CountryFolderFilePath + currentCountry)
        wb1.save(CountryFolderFilePath + currentCountry)
        wb1.save(CountryFolderFilePath + currentCountry)
        wb1.save(CountryFolderFilePath + currentCountry)
        wb1.save(CountryFolderFilePath + currentCountry)
        f = f + 1
AllFiles = GetAllFiles()
Main(AllFiles)