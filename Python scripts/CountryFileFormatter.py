import os
import os.path
import sys
from time import sleep
import webbrowser
import datetime
from datetime import time, date, timedelta
from openpyxl import Workbook, load_workbook, styles
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import xlrd 
import string
clear = lambda: os.system('cls')
clear()
totalFormatted = 0
#For Use In The GitHub Repository - https://github.com/OQ2000/COVID-19-By-Country
#Python script to be used for COVID-19-By-Country
#Developed by Owen Quinn

#This script was written to take a master data set that was previously made using all WHO(World Health Orginisation) Situation reports.
#This script will create an excel file respective for each new country it finds on the master sheet.

def GetAllFiles():
    AllFiles = os.listdir(r"C:\Users\gunsl_000\Desktop\COVID-19 By Country\COVID-19-By-Country\Datasets\SeperateCountryFiles")
    print("Found",len(AllFiles),"Files")
    print("Formatting All")
    sleep(2)
    return AllFiles

def ConvertNumberToLetter(CheckingNumber):
    NumberToLetter = {
    0 : "a",
    1 : "b",
    2 : "c",
    3 : "d",
    4 : "e",
    5 : "f",
    6 : "g",
    7 : "h",
    8 : "i",
    9 : "j",
    10 : "k",
    11 : "l",
    12 : "m",
    13 : "n",
    14 : "o",
    15 : "p",
    16 : "q",
    17 : "r",
    18 : "s",
    19 : "t",
    20 : "u",
    21 : "v",
    22 : "w",
    23 : "x",
    24 : "y",
    25 : "z"
    }
    return NumberToLetter[CheckingNumber]

def DateFormatter(CheckingCountry, wbook, sheet1, MaxRows):
    for cell in sheet1["E"]:
        cell.number_format = "yyyy-mm-dd;"

def GlobalFormatter(sheet1, sheet,MaxRows):
    #Sets Width
    for x in range(15):
        tmpX = ConvertNumberToLetter(x)
        sheet1.column_dimensions[tmpX].width = 27
        x = x + 1
        for y in range(1, MaxRows):
            sheet1.cell(row=y, column=x).alignment = Alignment(horizontal='center',vertical='center')
            y + y + 1

def Calulations(sheet, sheet1, MaxRows):
    #TotalConfirmedCases
    sheet.cell_value(0, 0) 
    for i in range(sheet.ncols):
        if sheet.cell_value(0, i) == "TotalConfirmedCases":
            TotalConfirmedCasesColumnRef = i
            TotalConfirmedCasesColumnRefLetter = ConvertNumberToLetter(TotalConfirmedCasesColumnRef)
    #Calulates New Cases
    TotalCases = []
    for cell in sheet1[TotalConfirmedCasesColumnRefLetter]:
        TotalCases.append(cell.value)
    #Adds Calulation
    tmpi = 0
    NewCasesRef = str(ConvertNumberToLetter(TotalConfirmedCasesColumnRef+1))
    for cell in sheet1[NewCasesRef]:
        if cell.row == 1:
            pass
        elif cell.row == 2:
            cell.value = TotalCases[1]
        else:
            cell.value = int(TotalCases[tmpi]) - int(TotalCases[tmpi-1])
        tmpi = tmpi +1
    ###############################
    #DeathCellRef
    sheet.cell_value(0, 0) 
    for i in range(sheet.ncols):
        if sheet.cell_value(0, i) == "TotalDeaths":
            DeathColumnRef = i
            DeathColumnRefLetter = ConvertNumberToLetter(DeathColumnRef)
    TotalDeaths = []
    for cell in sheet1[DeathColumnRefLetter]:
        TotalDeaths.append(cell.value)
    tmpx = 0
    NewDeathsRef = str(ConvertNumberToLetter(DeathColumnRef+1))
    for cell in sheet1[NewDeathsRef]:
        if cell.row == 1:
            pass
        elif cell.row == 2:
            cell.value = TotalDeaths[1]
        elif TotalDeaths[tmpx] == "N/R":
            cell.value = 0
        elif TotalDeaths[tmpx-1] == "N/R":
            cell.value = 0
        else:
            cell.value = int(TotalDeaths[tmpx]) - int(TotalDeaths[tmpx-1])
        tmpx = tmpx +1


def Main(totalFormatted, AllFiles):
    LenAllFiles = len(AllFiles)
    CountryFolderFilePath = r"C:\Users\gunsl_000\Desktop\COVID-19 By Country\COVID-19-By-Country\Datasets\SeperateCountryFiles/"
    for f in range(LenAllFiles):
        CheckingCountry = AllFiles[f]
        CheckingCountry = CheckingCountry[:-5]
        wb1 = load_workbook(CountryFolderFilePath + CheckingCountry + ".xlsx")
        wbook = xlrd.open_workbook(CountryFolderFilePath + CheckingCountry + ".xlsx")
        sheet = wbook.sheet_by_index(0)
        sheet1 = wb1[CheckingCountry]

        #Gets Max Rows for effeciency
        for i in range(sheet.nrows):
            i = i + 1
        MaxRows = i + 1
        DateFormatter(CheckingCountry, wbook, sheet1, MaxRows)
        GlobalFormatter(sheet1, sheet,MaxRows)
        Calulations(sheet, sheet1, MaxRows)
        wb1.save(CountryFolderFilePath + CheckingCountry + ".xlsx")
        wb1.save(CountryFolderFilePath + CheckingCountry + ".xlsx")
        print("Formatted:",CheckingCountry + ".xlsx")
        print("-------------------------------------------------------")
        totalFormatted = totalFormatted + 1
        f = f + 1
    return totalFormatted
        
        
def start_Program(totalFormatted):        
    AllFiles = GetAllFiles()
    totalFormatted = Main(totalFormatted,AllFiles)
    print("Finished Formatting, Formatted",totalFormatted)

start_Program(totalFormatted)