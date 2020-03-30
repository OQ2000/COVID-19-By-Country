from pynput.keyboard import Key, Controller, KeyCode, Listener
import clipboard
from time import sleep
import os
import os.path
import sys
import random
from pyautogui import press, hotkey, typewrite, keyDown, keyUp
import webbrowser
import datetime
from datetime import time, date, timedelta
from openpyxl import Workbook, load_workbook, styles
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import xlrd 
import string
#For Use In The GitHub Repository - https://github.com/OQ2000/COVID-19-By-Country
#Python script to be used for COVID-19-By-Country
#Developed by Owen Quinn

#This script was written to correctly format the Master-Data-Sheet for use with other scripts.
#It assigns each record with the correct situation report number and date for that report.

clear = lambda: os.system('cls') #on Windows System
clear()
CurrentRow = 2
MasterFilePath = r"C:\Users\gunsl_000\Desktop\COVID-19 By Country\COVID-19-By-Country\Datasets\Master-DATA-SET.xlsx"
wbook = xlrd.open_workbook(MasterFilePath)
sheet = wbook.sheet_by_index(0)
wb = load_workbook(MasterFilePath)
sh = wb["Datasets"]

def FixSitReps():
    LastCell = ""
    for cell in sh["a"]:
        if cell.value is None:
            print("Empty Cell At",cell.coordinate)
            sh[cell.coordinate].value = LastCell
            print("New Cell Data:",cell.value)
        else:
            print("Cell Not Empty, Contains",sh[cell.coordinate].value)
            LastCell = cell.value
            print(LastCell)
    wb.save(r"C:\Users\gunsl_000\Desktop\COVID-19 By Country\COVID-19-By-Country\Datasets\Master-DATA-SET.xlsx")
def FixDates():
    LastCell = ""
    for cell in sh["d"]:
        if cell.value is None:
            print("Empty Cell At",cell.coordinate)
            sh[cell.coordinate].value = LastCell
            print("New Cell Data:",cell.value)
        else:
            print("Cell Not Empty, Contains",sh[cell.coordinate].value)
            LastCell = cell.value
            print(LastCell)
    wb.save(r"C:\Users\gunsl_000\Desktop\COVID-19 By Country\COVID-19-By-Country\Datasets\Master-DATA-SET.xlsx")
    

FixSitReps()
FixDates()

