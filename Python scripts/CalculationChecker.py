import os
import os.path
import sys
from time import sleep
import webbrowser
import datetime
from datetime import time, date, timedelta
from openpyxl import Workbook, load_workbook, styles
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import xlrd 
import string
clear = lambda: os.system('cls')
clear()
totalIncorrect = 0
CountryFolderFilePath = r"C:\Users\gunsl_000\Desktop\COVID-19 By Country\COVID-19-By-Country\Datasets\SeperateCountryFiles/"

def GetAllFiles():
    AllFiles = os.listdir(r"C:\Users\gunsl_000\Desktop\COVID-19 By Country\COVID-19-By-Country\Datasets\SeperateCountryFiles")
    print("Found",len(AllFiles),"Files")
    print("Searching All For Errors")
    sleep(2)
    return AllFiles

def Main(AllFiles):
    problems = []
    lenAllFiles = len(AllFiles)
    for i in range(lenAllFiles):
        CheckingCountry = AllFiles[i]
        CheckingCountry = CheckingCountry[:-5]
        wb1 = load_workbook(CountryFolderFilePath + CheckingCountry + ".xlsx")
        wbook = xlrd.open_workbook(CountryFolderFilePath + CheckingCountry + ".xlsx")
        sheet = wbook.sheet_by_index(0)
        sheet1 = wb1[CheckingCountry]
        MaxRows = 0
        for i in range(sheet.nrows):
            i = i + 1
        MaxRows = i - 1
        currentRow = 2
        for t in range(1, MaxRows):
            if sheet1["H"+str(currentRow)].value  == "N/R":
                pass
            elif sheet1["H"+str(currentRow)].value < 0:
                problems.append(CheckingCountry)
                problems.append(sheet1["M"+str(currentRow)].value)
            if sheet1["J"+str(currentRow)].value == "N/R":
                pass
            elif sheet1["J"+str(currentRow)].value < 0:
                problems.append(CheckingCountry)
                problems.append(sheet1["M"+str(currentRow)].value)
            currentRow = currentRow + 1
        print("Checking",CheckingCountry)
    return problems

AllFiles = GetAllFiles()
problems = Main(AllFiles)

with open("Problems.txt", "w") as output:
    for i in range(0,len(problems)-1,2):
        line = str(problems[i]),"------RowNo.",str(problems[i+1])+"\n"
        output.writelines(line)

for i in range(0,len(problems)-1,2):
    print(problems[i],"\t",problems[i+1])