import os
import os.path
import sys
from time import sleep
import webbrowser
import datetime
from datetime import time, date, timedelta
from openpyxl import Workbook, load_workbook, styles
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import xlrd 
import string
clear = lambda: os.system('cls')
clear()
totalCorrected = 0
SeperateCountryFilePath = r"C:\Users\gunsl_000\Desktop\COVID-19 By Country\COVID-19-By-Country\Datasets\ToBeAddedCalculations/"

def GetAllFiles(SeperateCountryFilePath):
    AllFiles = os.listdir(SeperateCountryFilePath)
    print("Found",str(len(AllFiles)),"Files")
    print("Bringing All Over")
    return AllFiles

def GetRowNumber(sheet,sheet1,currentRow):
    tmpRowNum = 0
    tmpRowNum = sheet1["M"+str(currentRow)].value
    
    return tmpRowNum


def CopyingData(RowRef,CasesToCopy,DeathsToCopy,RowNum,sheet1,sheet):
    for i in range(sheet.nrows):
        i = i + 1
    MaxRows = i - 1
    currentRow = 2

    for i in range(MaxRows):
        RowRef.append(sheet1["M"+str(currentRow)].value)
        CasesToCopy.append(sheet1["H"+str(currentRow)].value)
        DeathsToCopy.append(sheet1["J"+str(currentRow)].value)
        currentRow = currentRow + 1
        i = i + 1
    return RowRef,CasesToCopy,DeathsToCopy


#For each file i need to take the row num and the two values in the H And J

def Main(AllFiles,SeperateCountryFilePath):
    LenAllFiles = len(AllFiles)
    RowRef = []
    CasesToCopy = []
    DeathsToCopy = []
    counter = 0
    for i in range(LenAllFiles):
        if counter == 5:
            counter = 0
            clear()
            for r in range(len(RowRef)):
                #Cases
                MasterSheet["H"+str(RowRef[r])].value = CasesToCopy[r]
                #Deaths
                MasterSheet["J"+str(RowRef[r])].value = DeathsToCopy[r]
                print("Copying To Cell","J"+str(RowRef[r]),"And","H"+str(RowRef[r]))
                print("Data:",CasesToCopy[r],"And",DeathsToCopy[r])
                r = r + 1
                print("Saving Master Document, This May Take A While")
            MasterWorkBook.save(MasterFile)
            RowRef = []
            CasesToCopy = []
            DeathsToCopy = []
        
        MasterFile = r"C:\Users\gunsl_000\Desktop\COVID-19 By Country\COVID-19-By-Country\Datasets\Master-DATA-SET.xlsx"
        CountryFolderFilePath = SeperateCountryFilePath
        MasterWorkBook = load_workbook(MasterFile)
        MasterSheet = MasterWorkBook["Datasets"]
        RowNum = 0
        CheckingCountry = AllFiles[i]
        CheckingCountry = CheckingCountry[:-5]
        print("----------------------------------")
        print(CheckingCountry)
        wb1 = load_workbook(CountryFolderFilePath + CheckingCountry + ".xlsx")
        wbook = xlrd.open_workbook(CountryFolderFilePath + CheckingCountry + ".xlsx")
        sheet = wbook.sheet_by_index(0)
        sheet1 = wb1[CheckingCountry]
        RowRef, CasesToCopy, DeathsToCopy = CopyingData(RowRef,CasesToCopy,DeathsToCopy,RowNum,sheet1,sheet)
        clear()
        print("All Vairables")
        print("Cases To Copy",CasesToCopy)
        print("Deaths To Copy",DeathsToCopy)
        print("RowRef",RowRef)
        counter = counter + 1

        # for r in range(MaxRows):
        #     RowNum = GetRowNumber(sheet,sheet1,currentRow)
        #     DataToCopy = CopyingData(currentRow,RowNum,sheet1,sheet)
        #     #Cases
        #     MasterSheet["H"+str(RowNum)].value = DataToCopy[0]
        #     #Deaths
        #     MasterSheet["J"+str(RowNum)].value = DataToCopy[1]
        #     currentRow = currentRow + 1
        #     r = r + 1
        # i = i + 1
    clear()
    for r in range(len(RowRef)):
        #Cases
        MasterSheet["H"+str(RowRef[r])].value = CasesToCopy[r]
        #Deaths
        MasterSheet["J"+str(RowRef[r])].value = DeathsToCopy[r]
        print("Copying To Cell","J"+str(RowRef[r]),"And","H"+str(RowRef[r]))
        print("Data:",CasesToCopy[r],"And",DeathsToCopy[r])
        r = r + 1
        print("Saving Master Document, This May Take A While")
    MasterWorkBook.save(MasterFile)
    # MasterWorkBook.save(MasterFile)
    
AllFiles = GetAllFiles(SeperateCountryFilePath)
Main(AllFiles,SeperateCountryFilePath)