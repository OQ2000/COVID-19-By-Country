from pynput.keyboard import Key, Controller, KeyCode, Listener
import clipboard
from time import sleep
import os
import os.path
import sys
import random
from pyautogui import press, hotkey, typewrite, keyDown, keyUp
import webbrowser
import datetime
from datetime import time, date, timedelta
from openpyxl import Workbook, load_workbook
import xlrd 
import string

CheckingCountryName = 'China'
wb = load_workbook(r"C:\Users\gunsl_000\Desktop\COVID-19 By Country\COVID-19-By-Country\TestingForMe\Testing.xlsx", data_only=True)
wb1 = load_workbook('TestingForMe\ContryTests'+'\\'+CheckingCountryName+'.xlsx')
sheet = wb["MainSheet"]
sheet1 = wb1["China"]
for row in sheet['A1':'Z1']:
    for cell in row:
        sheet1[cell.coordinate].value = cell.value
wb1.save('TestingForMe\ContryTests'+'\\'+CheckingCountryName+'.xlsx')