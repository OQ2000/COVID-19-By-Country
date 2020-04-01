from pynput.keyboard import Key, Controller, KeyCode, Listener
import clipboard
from time import sleep
import os
import os.path
import sys
import random
from pyautogui import press, hotkey, typewrite, keyDown, keyUp
import webbrowser
import datetime
from datetime import time, date, timedelta
from openpyxl import Workbook, load_workbook
import xlrd 
import string
def CountryFileCreater(CheckingCountryName):
    loc = (r"C:\Users\gunsl_000\Desktop\COVID-19 By Country\COVID-19-By-Country\TestingForMe\Testing.xlsx")
    wb = load_workbook(r"C:\Users\gunsl_000\Desktop\COVID-19 By Country\COVID-19-By-Country\TestingForMe\Testing.xlsx", data_only=True)
    wbook = xlrd.open_workbook(loc)
    sh = wb["MainSheet"]
    sheet = wbook.sheet_by_index(0)

    def ConvertNumberToLetter(CheckingNumber):
        NumberToLetter = {
        0 : "a",
        1 : "b",
        2 : "c",
        3 : "d",
        4 : "e",
        5 : "f",
        6 : "g",
        7 : "h",
        8 : "i",
        9 : "j",
        10 : "k",
        11 : "l",
        12 : "m",
        13 : "n",
        14 : "o",
        15 : "p",
        16 : "q",
        17 : "r",
        18 : "s",
        19 : "t",
        20 : "u",
        21 : "v",
        22 : "w",
        23 : "x",
        24 : "y",
        25 : "z"
        }
        return NumberToLetter[CheckingNumber]
    def CheckIfExcelFileForCountryExists(CheckingCountryName):
        CheckingCountryName = string.capwords(CheckingCountryName)
        CheckingCountryName = CheckingCountryName.replace(" ", "_")
        FilePath = 'TestingForMe\ContryTests'+'\\'+str(CheckingCountryName)+'.xlsx'
        if os.path.isfile('TestingForMe\ContryTests'+'\\'+str(CheckingCountryName)+'.xlsx'):
            print("File Exists")
        else:
            print("File Not Exist, Creating File")
            wb = Workbook()
            ws =  wb.active
            ws.title = CheckingCountryName
            wb.save(filename = 'TestingForMe\ContryTests'+'\\'+CheckingCountryName+'.xlsx')
        return FilePath
    FilePath = CheckIfExcelFileForCountryExists(CheckingCountryName)
    return FilePath
CheckingCountryName = "China"
FilePath = CountryFileCreater(CheckingCountryName)
print("File Path Is",FilePath)