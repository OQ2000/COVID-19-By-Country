from pynput.keyboard import Key, Controller, KeyCode, Listener
import clipboard
from time import sleep
import os
import os.path
import sys
import random
from pyautogui import press, hotkey, typewrite, keyDown, keyUp
import webbrowser
import datetime
from datetime import time, date, timedelta
from openpyxl import Workbook, load_workbook, styles
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import xlrd 
import string
clear = lambda: os.system('cls') #on Windows System
clear()

def ConvertNumberToLetter(CheckingNumber):
    NumberToLetter = {
    0 : "a",
    1 : "b",
    2 : "c",
    3 : "d",
    4 : "e",
    5 : "f",
    6 : "g",
    7 : "h",
    8 : "i",
    9 : "j",
    10 : "k",
    11 : "l",
    12 : "m",
    13 : "n",
    14 : "o",
    15 : "p",
    16 : "q",
    17 : "r",
    18 : "s",
    19 : "t",
    20 : "u",
    21 : "v",
    22 : "w",
    23 : "x",
    24 : "y",
    25 : "z"
    }
    return NumberToLetter[CheckingNumber]

def CheckIfExcelFileForCountryExists(MasterFilePath, CountryFolderFilePath, CheckingCountryName,wb):
    CheckingCountryName = string.capwords(CheckingCountryName)
    CheckingCountryName = CheckingCountryName.replace(" ", "_")
    CountryFilePath = CountryFolderFilePath+'\\'+str(CheckingCountryName)+'.xlsx'
    if os.path.isfile(CountryFilePath):
        print("File Exists")
    else:
        Masterwb = wb
        print("File Not Exist, Creating File")
        wb = Workbook()
        ws =  wb.active
        ws.title = CheckingCountryName
        wb.save(CountryFilePath)
        print(CountryFilePath)
        wb1 = load_workbook(CountryFilePath)
        sheet = Masterwb["MainSheet"]
        sheet1 = wb1[CheckingCountryName]
        for row in sheet['A1':'Z1']:
            for cell in row:
                sheet1[cell.coordinate].value = cell.value
        wb1.save(CountryFilePath)

def FindCountryNameColumn(MasterFilePath):
    loc = (MasterFilePath)
    wb = xlrd.open_workbook(loc)
    sheet = wb.sheet_by_index(0)
    # For row 0 and column 0 
    sheet.cell_value(0, 0) 
    for i in range(sheet.ncols):
        if sheet.cell_value(0, i) == "CountryName":
            CountryCellRefRow = i
            print('Country Column Is',CountryCellRefRow,"or",ConvertNumberToLetter(CountryCellRefRow))
            CountryCellRefRow = ConvertNumberToLetter(CountryCellRefRow)
            return CountryCellRefRow

def GettingNextLine(CurrentRow, MasterFilePath, CountryFolderFilePath, CheckingCountryName,CountryNameColumn,wb,wbook,sh,sheet):
    #Getting the max rows in the worksheet
    for i in range(sheet.nrows):
        i = i + 1
    MaxRows = i
    #Getting the next cell refernce to be copied
    for j in range(CurrentRow,MaxRows,1):
        j = j + 1
    CellRef = str(CountryNameColumn)+str(CurrentRow)
    CheckingCountryName = sh[CellRef].value
    return CheckingCountryName, CellRef

def CopyDataForCheckedCountryName(MasterFilePath,CountryFolderFilePath,CellRef,CheckingCountryName,TotalRecordsAdded,wb,wbook,sheet):
    for i in range(sheet.ncols):
        i = i + 1
    MaxRows = i - 1
    MaxRows = ConvertNumberToLetter(MaxRows)
    #CopyRowData]
    tmplength = len(CellRef)
    CellRef = CellRef[1:tmplength]
    CheckIfExcelFileForCountryExists(MasterFilePath, CountryFolderFilePath, CheckingCountryName,wb)
    CheckingCountryName = string.capwords(CheckingCountryName)
    CheckingCountryName = CheckingCountryName.replace(" ", "_")
    CountryFilePath = CountryFolderFilePath+'\\'+str(CheckingCountryName)+'.xlsx'
    wb1 = load_workbook(CountryFilePath)
    sheet = wb["MainSheet"]
    sheet1 = wb1[CheckingCountryName]
    StartCell = "a"+CellRef
    EndCell = MaxRows + CellRef
    # print("Start Cell Is:", StartCell)
    # print("End Cell Is:", EndCell)
    isDuplicate = False
    for cell in sheet1["a"]:
        if cell.value == sheet['a' + CellRef].value:
            print("Duplicate")
            isDuplicate = True
            break
        else:
            isDuplicate = False
    if isDuplicate == False:
        TotalRecordsAdded = TotalRecordsAdded + 1
        for cell in sheet1["e"]:
            if cell.value is None:
                NextEmptyCell = cell.row
                break
            else:
                NextEmptyCell = cell.row + 1
        print("Next Empty Cell In",CheckingCountryName,"Is",NextEmptyCell)
        for row in sheet[StartCell:EndCell]:
            for cell in row:
                tmpColumn = ConvertNumberToLetter(cell.column-1)
                sheet1[tmpColumn+str(NextEmptyCell)].value = cell.value
        wb1.save(CountryFilePath)
        return TotalRecordsAdded
    else:
        print("Skipping Duplicate Record")
        return TotalRecordsAdded
    #CopyData should use cell ref to get the row number and get all of that lines data.
    #Also should get max coloumn to get all data for a given row.

def UserInput():
    MasterFilePathFound = False
    while MasterFilePathFound == False:
        MasterFilePath = input("Please Enter The Directory Of The Master File: ")
        if os.path.isfile(MasterFilePath):
            clear()
            print("Searching For",MasterFilePath)
            print("File Found")
            MasterFilePathFound = True
        else:
            clear()
            print("Please Enter A Valid Directory(Include Excel Name + .xlsx At the End)")
    CountryFolderFilePathFound = False
    while CountryFolderFilePathFound == False:
        CountryFolderFilePath = input("Please Enter The Directory Of The Folder That The Country Files Will Be Saved Into: ")
        if os.path.exists(CountryFolderFilePath):
            clear()
            print("Searching For",CountryFolderFilePath)
            print("File Found")
            CountryFolderFilePathFound = True
        else:
            clear()
            print("Please Enter A Valid Directory")
    StartRowInt = False
    while StartRowInt == False:
        StartRow = input("Enter The Row That You Want To Start Copying From: ")
        if (StartRow.isdigit()):
            StartRowInt = True
        else:
            clear()
            print("Entry Is Not An Integer")
    return MasterFilePath, CountryFolderFilePath, StartRow

def Start_Program():
    #Default False Used For Testing
    TotalRecordsAdded = 0
    tmpAnswered = True
    if tmpAnswered == False:
        MasterFilePath, CountryFolderFilePath, StartRow = UserInput()
        clear()
        CurrentRow = StartRow
        while tmpAnswered == False:
            clear()
            print("The MasterFilePath Is:",MasterFilePath)
            print("The CountryFolderFilePath Is:",CountryFolderFilePath)
            print("The StartRow Is:",CurrentRow)
            tmpAnswer = input("Are These Correct? Y/N")
            if tmpAnswer == "N":
                clear()
                UserInput()
            elif tmpAnswer == "Y":
                print("user selected Yes")
                tmpAnswered = True
            else:
                print("Please Enter Y or N: ")
                clear()
    
    else:
        CurrentRow = 1
        if CurrentRow == 0 or CurrentRow == 1:
            CurrentRow = 2
        MasterFilePath = r"C:\Users\gunsl_000\Desktop\COVID-19 By Country\COVID-19-By-Country\TestingForMe\Testing.xlsx"
        CountryFolderFilePath = r"C:\Users\gunsl_000\Desktop\COVID-19 By Country\COVID-19-By-Country\TestingForMe\CountryTest"
        CheckingCountryName = ""
        wbook = xlrd.open_workbook(MasterFilePath)
        sheet = wbook.sheet_by_index(0)
        wb = load_workbook(MasterFilePath, data_only=True)
        sh = wb["MainSheet"]
        sheet = wbook.sheet_by_index(0)
        for i in range(sheet.nrows):
            i = i + 1
        MaxRows = i - 1
        print("Max Rows:",MaxRows)
        CountryNameColumn = FindCountryNameColumn(MasterFilePath)
        for i in range(MaxRows):
            CheckingCountryName, CellRef = GettingNextLine(CurrentRow, MasterFilePath, CountryFolderFilePath, CheckingCountryName, CountryNameColumn,wb,wbook,sh,sheet)
            print(CheckingCountryName)
            print("Checking:",CellRef)
            TotalRecordsAdded = CopyDataForCheckedCountryName(MasterFilePath,CountryFolderFilePath,CellRef,CheckingCountryName,TotalRecordsAdded,wb,wbook,sheet)
            CurrentRow = CurrentRow + 1
            i = i + 1
    print("Added",TotalRecordsAdded,"Records This Session")

Start_Program()

